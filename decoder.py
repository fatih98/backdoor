from openpyxl import Workbook,load_workbook
x=0
list = []
text = ''
realText = ''
file = open("deneme.txt")
message = file.read()
message = message.split('&')
wb = load_workbook("deneme.xlsx")

while(x<len(message)):
	text = message[x]
	list += text.split('!')
	x+=1
x=0
try:
	while(x<=len(list)*2):
		print()
		ws = wb[list[x]]
		x+=1
		realText += ws[list[x]].value
		x+=1
except:
	print("")
print(realText)